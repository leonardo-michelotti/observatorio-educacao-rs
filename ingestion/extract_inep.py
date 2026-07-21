#!/usr/bin/env python3
"""Planilhas oficiais do Inep -> Parquet bronze de rendimento e TDI.

Descobre os arquivos ZIP nas páginas anuais oficiais, mantém uma cópia em cache e lê somente
Brasil, Rio Grande do Sul e Santa Maria. O parser aceita os formatos históricos XLS e os XLSX
atuais e usa os cabeçalhos editoriais, não os nomes técnicos que mudam entre edições.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import time
import unicodedata
import warnings
import zipfile
from datetime import UTC, datetime
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Iterable, Iterator, Sequence

import certifi
import pandas as pd
import requests
import xlrd
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

ROOT = Path(__file__).resolve().parents[1]
BRONZE = ROOT / "data" / "bronze"
CACHE = ROOT / "data" / "raw" / "inep"
SOURCES_PATH = Path(__file__).with_name("inep_sources.json")
REFERENCE_PATH = Path(__file__).with_name("inep_reference_values.json")
INEP_INTERMEDIATE = Path(__file__).with_name("certs") / "rnp-icpedu-gr46-ov-tls-ca-2025.pem"
MUNICIPIO_SANTA_MARIA = 4316907
LEVELS = ("brasil", "rs", "santa_maria")


class _Links(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.urls: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag != "a":
            return
        href = dict(attrs).get("href") or ""
        if "download.inep.gov.br" in href and href.lower().endswith(".zip"):
            self.urls.append(href)


def _normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().casefold()


def _year(value: object) -> int | None:
    try:
        number = int(float(value))
    except (TypeError, ValueError):
        return None
    return number if 2000 <= number <= 2100 else None


def _number(value: object) -> float | None:
    if value is None or _normalize(value) in {"", "--", "-", "nd", "n/d"}:
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def _column_text(header: Sequence[Sequence[object]], index: int) -> str:
    return " | ".join(_normalize(row[index]) for row in header if index < len(row) and row[index])


def _stage_columns(header: Sequence[Sequence[object]], source: str) -> dict[str, int]:
    width = max(map(len, header), default=0)
    texts = {index: _column_text(header, index) for index in range(width)}
    approval_start = 0
    approval_end = width
    if source == "rendimento":
        starts = [
            index
            for row in header
            for index, value in enumerate(row)
            if "aprov" in _normalize(value)
        ]
        if starts:
            approval_start = min(starts)
        ends = [
            index
            for row in header
            for index, value in enumerate(row)
            if index > approval_start
            and ("reprov" in _normalize(value) or "abandono" in _normalize(value))
        ]
        if ends:
            approval_end = min(ends)

    def find(predicate, label: str) -> int:
        matches = [index for index, text in texts.items() if predicate(index, text)]
        if len(matches) != 1:
            raise ValueError(f"Cabeçalho {source}: coluna {label} ambígua ou ausente: {matches}")
        return matches[0]

    def correct_family(index: int, text: str) -> bool:
        if source == "tdi":
            return True
        return approval_start <= index < approval_end or "1_cat" in text or "tap_" in text

    initial = find(
        lambda index, text: correct_family(index, text)
        and (
            "anos iniciais" in text
            or bool(re.search(r"1\D+(?:a|ao)\D+(?:4|5)\D+(?:serie|ano)", text))
        ),
        "EF anos iniciais",
    )
    final = find(
        lambda index, text: correct_family(index, text)
        and (
            "anos finais" in text
            or bool(re.search(r"5\D+(?:a|ao)\D+8\D+serie|6\D+ao\D+9\D+ano", text))
        ),
        "EF anos finais",
    )
    high_school = find(
        lambda index, text: correct_family(index, text)
        and "medio" in text
        and "total" in text
        and not re.search(r"nao[- ]seriado", text),
        "Ensino Médio",
    )
    return {"ef_anos_iniciais": initial, "ef_anos_finais": final, "em": high_school}


def _is_municipality_code(value: object) -> bool:
    try:
        number = int(float(value))
    except (TypeError, ValueError):
        return False
    return 1_000_000 <= number <= 9_999_999


def _row_level(
    row: Sequence[object], value_start: int, default_level: str | None = None
) -> str | None:
    metadata = list(row[:value_start])
    normalized = [_normalize(value) for value in metadata]
    if normalized.count("total") < 2:
        return None
    if any(
        _is_municipality_code(value) and int(float(value)) == MUNICIPIO_SANTA_MARIA
        for value in metadata
    ):
        return "santa_maria"
    if "brasil" in normalized:
        return "brasil"
    has_municipality_code = any(_is_municipality_code(value) for value in metadata)
    if not has_municipality_code and ({"rs", "rio grande do sul"} & set(normalized)):
        return "rs"
    if not has_municipality_code and default_level:
        return default_level
    return None


def _parse_rows(
    rows: Iterable[Sequence[object]],
    source: str,
    expected_year: int,
    wanted: set[str] | None = None,
    default_level: str | None = None,
) -> dict[str, dict]:
    header: list[Sequence[object]] = []
    columns: dict[str, int] | None = None
    found: dict[str, dict] = {}
    prefixes = {"rendimento": "taxa_aprovacao", "tdi": "tdi"}

    for row in rows:
        row_year = _year(row[0] if row else None)
        if row_year is None:
            if columns is None:
                header.append(row)
            continue
        if columns is None:
            columns = _stage_columns(header, source)
        if row_year != expected_year:
            continue
        level = _row_level(row, min(columns.values()), default_level)
        if level is None:
            continue
        values = {
            f"{prefixes[source]}_{stage}": _number(row[index] if index < len(row) else None)
            for stage, index in columns.items()
        }
        if level in found and found[level] != values:
            raise ValueError(f"Mais de uma linha candidata para {source}/{expected_year}/{level}")
        found[level] = values
        if wanted and wanted <= set(found):
            break
    return found


def _xlsx_sheets(content: bytes) -> Iterator[Iterable[Sequence[object]]]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        yield sheet.iter_rows(values_only=True)


def _xls_sheets(content: bytes) -> Iterator[Iterable[Sequence[object]]]:
    workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
    for sheet in workbook.sheets():
        yield (sheet.row_values(index) for index in range(sheet.nrows))


def parse_archive(
    path: Path, source: str, year: int, wanted: set[str] | None = None
) -> dict[str, dict]:
    found: dict[str, dict] = {}
    with zipfile.ZipFile(path) as archive:
        spreadsheets = [
            name for name in archive.namelist() if Path(name).suffix.lower() in {".xls", ".xlsx"}
        ]
        if not spreadsheets:
            raise ValueError(f"Nenhuma planilha encontrada em {path}")
        for name in spreadsheets:
            content = archive.read(name)
            sheets = _xlsx_sheets(content) if name.lower().endswith(".xlsx") else _xls_sheets(content)
            filename = _normalize(Path(name).stem)
            national_only = (
                "brasil" in filename
                and "regio" not in filename
                and not re.search(r"(?:^|\W)ufs?(?:\W|$)", filename)
            )
            default_level = "brasil" if national_only else None
            for rows in sheets:
                for level, values in _parse_rows(
                    rows, source, year, wanted, default_level
                ).items():
                    if level in found and found[level] != values:
                        raise ValueError(f"Conflito em {path.name}: {level}")
                    found[level] = values
                if wanted and wanted <= set(found):
                    return found
    return found


def _ca_bundle() -> Path:
    """Combina raízes Mozilla com o intermediário que o host legado do Inep não envia."""
    destination = CACHE / "ca-bundle.pem"
    content = Path(certifi.where()).read_bytes() + b"\n" + INEP_INTERMEDIATE.read_bytes()
    if not destination.exists() or destination.read_bytes() != content:
        destination.parent.mkdir(parents=True, exist_ok=True)
        destination.write_bytes(content)
    return destination


def _session() -> requests.Session:
    retry = Retry(
        total=3,
        connect=3,
        read=3,
        backoff_factor=1,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
    )
    session = requests.Session()
    session.headers["User-Agent"] = "observatorio-educacao-rs/1.0 (dados publicos)"
    session.verify = str(_ca_bundle())
    session.mount("https://", HTTPAdapter(max_retries=retry))
    return session


def discover_urls(session: requests.Session, page: str) -> dict[str, str]:
    response = session.get(page, timeout=60)
    response.raise_for_status()
    parser = _Links()
    parser.feed(response.text)
    urls = list(dict.fromkeys(parser.urls))
    selected = {
        "br_uf": [url for url in urls if "brasil" in _normalize(url) and "escola" not in _normalize(url)],
        "municipios": [url for url in urls if "municip" in _normalize(url)],
    }
    if any(len(matches) != 1 for matches in selected.values()):
        raise ValueError(f"Links oficiais inesperados em {page}: {selected}")
    return {scope: matches[0] for scope, matches in selected.items()}


def download(session: requests.Session, url: str, destination: Path) -> Path:
    if destination.exists() and zipfile.is_zipfile(destination):
        return destination
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(".part")
    for attempt in range(1, 6):
        try:
            with session.get(url, timeout=(30, 120), stream=True) as response:
                response.raise_for_status()
                with temporary.open("wb") as output:
                    for chunk in response.iter_content(1024 * 1024):
                        output.write(chunk)
            if not zipfile.is_zipfile(temporary):
                raise ValueError("resposta não é um ZIP")
            temporary.replace(destination)
            return destination
        except (OSError, requests.RequestException, ValueError):
            temporary.unlink(missing_ok=True)
            if attempt == 5:
                raise
            time.sleep(attempt * 2)
    raise RuntimeError("download sem retorno")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _requested_years(specification: str | None, minimum: int, maximum: int) -> range:
    if specification is None:
        return range(minimum, maximum + 1)
    start, separator, end = specification.partition(":")
    first = int(start)
    last = int(end) if separator else first
    if first < minimum or last > maximum or first > last:
        raise ValueError(f"Anos fora do intervalo {minimum}:{maximum}: {specification}")
    return range(first, last + 1)


def extract(years: str | None = None) -> tuple[pd.DataFrame, list[dict]]:
    sources = json.loads(SOURCES_PATH.read_text(encoding="utf-8"))
    session = _session()
    records: dict[tuple[int, str], dict] = {}
    provenance: list[dict] = []

    for source, config in sources.items():
        minimum, maximum = config["years"]
        for year in _requested_years(years, minimum, maximum):
            page = config["page"].format(year=year)
            urls = discover_urls(session, page)
            year_values: dict[str, dict] = {}
            for scope, url in urls.items():
                archive = download(session, url, CACHE / source / str(year) / f"{scope}.zip")
                wanted = {"brasil", "rs"} if scope == "br_uf" else {"santa_maria"}
                parsed = parse_archive(archive, source, year, wanted)
                year_values.update(parsed)
                provenance.append(
                    {
                        "source": source,
                        "year": year,
                        "scope": scope,
                        "url": url,
                        "sha256": _sha256(archive),
                        "bytes": archive.stat().st_size,
                    }
                )
            missing = set(LEVELS) - set(year_values)
            if missing:
                raise ValueError(f"Ausentes em {source}/{year}: {sorted(missing)}")
            for level, values in year_values.items():
                records.setdefault((year, level), {"nivel": level, "ano": year}).update(values)
            print(f"  {source} {year}: Brasil, RS e Santa Maria")

    frame = pd.DataFrame(records.values()).sort_values(["ano", "nivel"]).reset_index(drop=True)
    return frame, provenance


def validate_reference(frame: pd.DataFrame) -> None:
    reference = json.loads(REFERENCE_PATH.read_text(encoding="utf-8"))
    for year_text, levels in reference.items():
        year = int(year_text)
        if year not in set(frame["ano"]):
            continue
        for level, expected in levels.items():
            row = frame[(frame["ano"] == year) & (frame["nivel"] == level)]
            if len(row) != 1:
                raise ValueError(f"Referência sem linha única: {year}/{level}")
            for column, value in expected.items():
                actual = row.iloc[0][column]
                if pd.isna(actual) or abs(float(actual) - value) > 1e-9:
                    raise ValueError(
                        f"Referência divergente {year}/{level}/{column}: {actual} != {value}"
                    )


def merge_existing(frame: pd.DataFrame, path: Path) -> pd.DataFrame:
    if not path.exists():
        return frame
    existing = pd.read_parquet(path)
    refreshed_years = set(frame["ano"])
    historical = existing[~existing["ano"].isin(refreshed_years)]
    return (
        pd.concat([historical, frame], ignore_index=True, sort=False)
        .sort_values(["ano", "nivel"])
        .reset_index(drop=True)
    )


def latest_year() -> int:
    sources = json.loads(SOURCES_PATH.read_text(encoding="utf-8"))
    years = {config["years"][1] for config in sources.values()}
    if len(years) != 1:
        raise ValueError(f"Fontes com anos máximos diferentes: {sorted(years)}")
    return years.pop()


def main() -> None:
    parser = argparse.ArgumentParser()
    years = parser.add_mutually_exclusive_group()
    years.add_argument("--years", help="ano único ou intervalo inclusivo, por exemplo 2023:2025")
    years.add_argument("--latest", action="store_true", help="baixa somente o último ano configurado")
    parser.add_argument("--merge", action="store_true", help="substitui os anos baixados no bronze existente")
    args = parser.parse_args()

    specification = str(latest_year()) if args.latest else args.years
    frame, provenance = extract(specification)
    validate_reference(frame)
    BRONZE.mkdir(parents=True, exist_ok=True)
    output = BRONZE / "indicadores.parquet"
    if args.merge:
        frame = merge_existing(frame, output)
    frame.to_parquet(output, index=False)
    metadata = {
        "generated_at": datetime.now(UTC).isoformat(),
        "municipality": MUNICIPIO_SANTA_MARIA,
        "mode": "merge" if args.merge else "replace",
        "files": provenance,
    }
    (BRONZE / "inep_provenance.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(f"Indicadores oficiais do Inep: {len(frame)} linhas em {BRONZE}")


if __name__ == "__main__":
    main()
